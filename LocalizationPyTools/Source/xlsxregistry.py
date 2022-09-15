# author: Katarzyna 'K8' Sosnowska
# e-mail: sosnowska.kk@gmail.com
# 
# date of last update: 2022-09-15
# 
# # About: 
# This script contains method for generating one XLSX file with a record of all texts merged from regular XLSX translation files.
#
# The generated record does not contains any translations for given texts. It only specifies where are translations for each text
# that exists in one of input XLSX files.
# The reason for creating this tool was the need to speed up looking up where are translations for a particular localized text.
# If you need to edit/remove some translations, you can quickly check in which XLSX file those translations are.
# 
# .

import os
from configtools import LocToolsConfig
from datetime import datetime
from openpyxl import Workbook
from openpyxl import load_workbook  # for working with xlsx files

STR_TEXT_ID = "id"
STR_COMMENTS = "comments"
STR_NONE = "none"

def generate_translations_registry(_input_dir_path, _output_dir_path, _debug_on):
    if not os.path.exists(_input_dir_path):
        print("[xlsxregistry] ERROR: \"{0}\" directory doesn't exist!".format(_input_dir_path))
        return {}
    # analyse translations from all XLSX files:
    if _debug_on: print("[xlsxregistry] : analysing XLSX files in \"{0}\" directory".format(_input_dir_path))
    debug_analysed_texts_num = 0
    translations_list = []
    for input_file_name in os.listdir(_input_dir_path):
        if not input_file_name.endswith(".xlsx"):
            continue
        input_file_path = os.path.join(_input_dir_path, input_file_name)
        if _debug_on: print("  > analysing \"{0}\" file...".format(input_file_name))
        wb = load_workbook(input_file_path)
        ws = wb.active
        # cache letters of columns with text id, source string and comments:
        text_id_column = "A"
        comments_column = "B"
        source_string_column = "C"
        for col in ws.iter_cols(min_row=1, min_col=1, max_col=10):
            cell = col[0]
            if cell.value is None:
                continue
            cell_str = str(cell.value).lower()
            if cell_str == "id":
                text_id_column = cell.column_letter
                continue
            elif cell_str == "en":
                source_string_column = cell.column_letter
                continue
            elif cell_str == "context":
                comments_column = cell.column_letter
                continue
            pass
        # record all found translations:
        num_of_elements = len(ws[text_id_column])
        for row_index in range(2, num_of_elements):
            text_id_raw = ws['{0}{1}'.format(text_id_column, row_index)].value
            if text_id_raw is None:
                continue
            text_id_elements = text_id_raw.split('.csv+')
            if len(text_id_elements) < 2:
                continue
            text_id = text_id_elements[1].rstrip('\r\n')
            source_string = ws['{0}{1}'.format(source_string_column, row_index)].value
            text_comments = ws['{0}{1}'.format(comments_column, row_index)].value
            text_source_file = text_id_elements[0]
            # add new record to the translations registry:
            translations_list.append((text_id, source_string, text_comments, text_source_file, input_file_name))
            debug_analysed_texts_num += 1
            pass
    if _debug_on: print("[xlsxregistry] : analysed {0} texts (in total)".format(debug_analysed_texts_num))
    # create a XLSX output file:
    wb = Workbook()
    ws = wb.active
    ws.title = "All existing translations"
    ws["A1"] = "ID"
    ws["B1"] = "Source Text"
    ws["C1"] = "Context"
    ws["D1"] = "Text Source File"
    ws["E1"] = "Translations Batch"
    index = 2
    for entry in translations_list:
        ws["A{0}".format(index)] = entry[0]
        ws["B{0}".format(index)] = entry[1]
        ws["C{0}".format(index)] = entry[2]
        ws["D{0}".format(index)] = entry[3]
        ws["E{0}".format(index)] = entry[4]
        index += 1
    # save results to the new output file:
    if not os.path.exists(_output_dir_path):
        os.mkdir(_output_dir_path)
    output_file_name = "TranslationsRegistry_{0}.xlsx".format(datetime.now().strftime("%Y%m%d%H%M%S"))
    output_path = os.path.join(_output_dir_path, output_file_name)
    wb.save(output_path)
    pass

#===============================================================================================

if __name__== "__main__":
    config_data = LocToolsConfig()
    config_data.load_config_data()
    xlsxan_settings = config_data.get_xlsxreg_settings()
    generate_translations_registry(xlsxan_settings[0], xlsxan_settings[1], True)
