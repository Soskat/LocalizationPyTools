# author: Katarzyna 'K8' Sosnowska
# e-mail: sosnowska.kk@gmail.com
# 
# date of last update: 2022-08-21
# 
# # About: 
# This file contains methods for generating a list of source texts that have changed.
# 
# # How LCT program works?
#
# This simple program analyses data from the input PO file with native culture.
# It compares the content of the `msgstr` and the `msgid` parts of the localized text data.
# If `msgstr` text is different than the text assigned to `msgid`, it means that the source text in `msgid`
# has changed, and the existing translation stored in `msgstr` property may be outdated.
#
# .

import os
from configtools import LocToolsConfig
from datetime import datetime
from openpyxl import Workbook

TEXT_KEY_TAG = "#. Key:\t"
TEXT_SOURCE_LOCATION_TAG = "#. SourceLocation:\t"
TEXT_COMMENTS_TAG = "#. InfoMetaData:\t\"Comments\" : "
MSGID = "msgid "        # source text
MSGSTR = "msgstr "      # translation text

'''
Creates a new workbook and populates it with entries from collected list of changed texts.
'''
def create_new_workbook(changed_texts):
    wb = Workbook()
    ws = wb.active
    ws.title = "Changed texts"
    ws["A1"] = "ID"
    ws["B1"] = "Source Location"
    ws["C1"] = "Comments"
    ws["D1"] = "Old text"
    ws["E1"] = "Current source text"
    row_index = 2
    for index in range(0, len(changed_texts) - 1):
        text_key, source_location, info_comments, new_text, old_text = changed_texts[index]
        row_index = index + 2
        ws["A{0}".format(row_index)] = text_key
        ws["B{0}".format(row_index)] = source_location
        ws["C{0}".format(row_index)] = info_comments
        ws["D{0}".format(row_index)] = old_text
        ws["E{0}".format(row_index)] = new_text
    return wb

'''
Analyses PO file with native languages and lists all text entries, that have differences between their MSGID and MSGSTR values.
All texts that will be on that list are the ones with chaned source text.
'''
def list_changed_native_texts(input_po_file_path, output_xlsx_dir_path):
    changed_texts = []
    debug_texts_count = 0
    # go line by line and check if there are any differences between source text and its translation:
    with open(input_po_file_path, "r", encoding="utf8") as input_file:
        text_key = ""
        source_location = ""
        info_comments = ""
        source_text = ""
        translation_text = ""
        for line in input_file:
            if line.startswith(TEXT_KEY_TAG):
                text_key = line[len(TEXT_KEY_TAG):-1]
                debug_texts_count += 1
            elif line.startswith(TEXT_SOURCE_LOCATION_TAG):
                source_location = line[len(TEXT_SOURCE_LOCATION_TAG):-1]
            elif line.startswith(TEXT_COMMENTS_TAG):
                info_comments = line[len(TEXT_COMMENTS_TAG) + 1:-2]
            elif line.startswith(MSGID):
                source_text = line[len(MSGID) + 1:-2]
            elif line.startswith(MSGSTR):
                translation_text = line[len(MSGSTR) + 1:-2]
                if source_text != translation_text:
                    changed_texts.append((text_key, source_location, info_comments, source_text, translation_text))
        pass
    # save list of changed texts to the output file:
    wb = create_new_workbook(changed_texts)
    ws = wb.active
    if not os.path.exists(output_xlsx_dir_path):
        os.mkdir(output_xlsx_dir_path)
    output_file_name = "ChangedTexts_{0}.xlsx".format(datetime.now().strftime("%Y%m%d%H%M%S"))
    output_path = os.path.join(output_xlsx_dir_path, output_file_name)
    wb.save(output_path)
    # log summary:
    print("\t> analysed \"{0}\" file [{1} texts | {2} changed translations]".format(input_po_file_path, debug_texts_count, len(changed_texts)))
    pass

#===============================================================================================

if __name__== "__main__":
    config_data = LocToolsConfig()
    config_data.load_config_data()
    lct_settings = config_data.get_lct_settings()
    list_changed_native_texts(lct_settings[0], lct_settings[1])
