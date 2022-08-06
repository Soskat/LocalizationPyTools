# author: Katarzyna 'K8' Sosnowska
# e-mail: sosnowska.kk@gmail.com
# 
# date of last update: 2022-08-06
# 
# # About: 
# This file contains methods for analysing translations listed in XLSX files iside specified directory.
# 
# # Base assumptions for this pipeline are as follows:
# - all string tables used in UE project are created from data loaded from CSV files
# - all localization data have been exported from UE built-in Localization system to a bunch of PO files
# - translation company provided a set of XLSX files with texts translations
#
# .

import os
from configtools import LocToolsConfig
from openpyxl import load_workbook  # for working with xlsx files

STR_TEXT_ID = "id"
STR_COMMENTS = "comments"
'''
Analyses texts translations from all input file from a given directory.
Param: inputDirPath - a path to a directory with input XLSX files that contain translations data.
Returns: generated dictionary with translations.
'''
def analyse_xlsx_translations(_input_dir_path, _max_column_number, _debug_on):
    if not os.path.exists(_input_dir_path):
        print("[xlsxanalyser] ERROR: \"{0}\" directory doesn't exist!".format(_input_dir_path))
        return {}
    # analyse translations from all XLSX files:
    if _debug_on: print("[xlsxanalyser] : analysing XLSX files in \"{0}\" directory".format(_input_dir_path))
    debug_analysed_texts_num = 0
    column_headers_map = {}
    translations_dic = {}
    for input_file_name in os.listdir(_input_dir_path):
        if not input_file_name.endswith(".xlsx"):
            continue
        input_file_path = os.path.join(_input_dir_path, input_file_name)
        if _debug_on: print("  > analysing \"{0}\" file...".format(input_file_name))
        wb = load_workbook(input_file_path)
        ws = wb.active
        # detect columns content:
        ids_column = "A"
        for col in ws.iter_cols(min_row=1, min_col=1, max_col=_max_column_number):
            cell = col[0]
            if cell.value is None:
                continue
            cell_str = str(cell.value).lower()
            if cell_str == STR_TEXT_ID:
                ids_column = cell.column_letter
                continue
            elif cell_str == STR_COMMENTS:
                continue
            column_headers_map[cell.column_letter] = cell.value
            if _debug_on: print("\t> found culture: \"{0}\"-\"{1}\"".format(cell.column_letter, cell.value))
            pass
        # analyse all found translation data:
        num_of_elements = len(ws[ids_column])
        for row_index in range(2, num_of_elements):
            text_id_raw = ws['{0}{1}'.format(ids_column, row_index)].value
            if text_id_raw is None:
                continue
            text_id_elements = text_id_raw.split('.csv+')
            if len(text_id_elements) < 2:
                continue
            text_id = text_id_elements[1].rstrip('\r\n')
            if text_id not in translations_dic:
                translations_dic[text_id] = {}
            for letter, culture_code in column_headers_map.items():
                cell = ws['{0}{1}'.format(letter, row_index)]
                if cell.value is None:
                    continue
                translations_dic[text_id][culture_code] = cell.value
            debug_analysed_texts_num += 1
            pass
    if _debug_on: print("[xlsxanalyser] : analysed {0} texts (in total)".format(debug_analysed_texts_num))
    return column_headers_map, translations_dic

#===============================================================================================

if __name__== "__main__":
    config_data = LocToolsConfig()
    config_data.load_config_data()
    xlsxan_settings = config_data.get_xlsxan_settings()
    column_headers_map, translations_dict = analyse_xlsx_translations(xlsxan_settings[0], xlsxan_settings[1], True)
