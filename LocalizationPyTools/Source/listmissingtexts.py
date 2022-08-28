# author: Katarzyna 'K8' Sosnowska
# e-mail: sosnowska.kk@gmail.com
# 
# date of last update: 2022-08-06
# 
# # About: 
# This file contains methods for generating a list of texts with missing translations.
# 
# # Base assumptions for this pipeline are as follows:
# - all string tables used in UE project are created from data loaded from CSV files
# - all localization data have been exported from UE built-in Localization system to a bunch of PO files
# - translation company provided a set of XLSX files with texts translations
# - texts IDs in XLSX files are composed with CSV file name and text key used in UE separated by the '+' sign:
#   eg.: 'ST_SomeFile.csv+Some_Text_ID'
# 
# # How LMT program works?
# 
# First, it analyses all XLSX files with translations and register all translations in a dedicated dictionary.
# For this step, the xlsxanalyser tool is used.
# 
# After that, the program goes through all input CSV files (those files are used to create string tables used in UE project) 
# and read their content. Each text ID from CSV file that is not in generated translations dictionary is written down 
# into output XLSX file. As a result, the output XLSX file will contain all texts that are missing their translations.
#
# .

import os
from configtools import LocToolsConfig
from datetime import datetime
from openpyxl import Workbook
from xlsxanalyser import analyse_xlsx_translations

XLSX_FILE_EXT = ".xlsx"
CSV_EXT = '.csv'

'''
Helper method that creates a new workbook and sets the name of initial sheet to given one.
'''
def create_new_workbook(sheet_name, column_headers_map):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for letter, cell_value in column_headers_map.items():
        ws["{0}1".format(letter)] = cell_value
    return wb

'''
Generates a list of texts that can't be found in provided XLSX files with translations.
All texts that will be on that list are missing translations.
'''
def list_missing_texts(csv_source_dir, xlsx_translations_dir, xlsx_output_dir, separator):
    if not os.path.exists(csv_source_dir):
        print("[listmissingtexts] ERROR: {0} directory doesn't exist!".format(csv_source_dir))
        return
    if not os.path.exists(xlsx_translations_dir):
        print("[listmissingtexts] ERROR: {0} directory doesn't exist!".format(xlsx_translations_dir))
        return
    column_headers_map, translations_dict = analyse_xlsx_translations(xlsx_translations_dir, 20, False)
    wb = create_new_workbook('Missing translations', column_headers_map)
    ws = wb.active
    # some stats:
    debug_total_text_count = 0
    debug_total_missing_text_count = 0
    # go through all source CSV files and check if they exist in translationsDict;
    # if not, then log such texts in output XLSX file:
    row_index = 2
    for file_name in os.listdir(csv_source_dir):
        debug_texts_count = 0
        debug_missing_texts_count = 0
        if file_name.endswith(CSV_EXT):
            file_full_name = os.path.join(csv_source_dir, file_name)
            input_file = open(file_full_name, "r", encoding='utf8')
            for line in input_file.readlines():
                words = line.split(separator)
                if len(words) < 2:
                    continue
                elif words[0].lower() == "\"key" or len(words[0]) < 1:
                    continue
                text_id = words[0][1:].rstrip('\"\n')
                if text_id not in translations_dict:
                    ws['A{0}'.format(row_index)].value = "{0}+{1}".format(file_name, text_id)
                    if len(words) > 2:
                        ws['C{0}'.format(row_index)].value = words[1].rstrip('\"\n').lstrip('\"')
                        ws['B{0}'.format(row_index)].value = words[2].rstrip('\"\n').lstrip('\"')
                    else:
                        ws['C{0}'.format(row_index)].value = words[1].rstrip('\",\n').lstrip('\"')
                    row_index += 1
                    debug_missing_texts_count += 1
                debug_texts_count += 1
            input_file.close()
            print("{0} [{1} texts checked | {2} missing texts]".format(file_name, debug_texts_count, debug_missing_texts_count))
            debug_total_text_count += debug_texts_count
            debug_total_missing_text_count += debug_missing_texts_count
        else:
            continue
    # add last one line with EOF ending:
    for letter in column_headers_map.keys():
        ws["{0}{1}".format(letter, row_index)] = "EOF"
    # save results to the new output file:
    if not os.path.exists(xlsx_output_dir):
        os.mkdir(xlsx_output_dir)
    output_file_name = "MissingTexts_{0}.xlsx".format(datetime.now().strftime("%Y%m%d%H%M%S"))
    output_path = os.path.join(xlsx_output_dir, output_file_name)
    wb.save(output_path)
    print("\nSUMMARY: {0} total texts checked | {1} total missing texts\n".format(debug_total_text_count, debug_total_missing_text_count))
    pass

#===============================================================================================

if __name__== "__main__":
    config_data = LocToolsConfig()
    config_data.load_config_data()
    xlsxan_settings = config_data.get_xlsxan_settings()
    ltm_settings = config_data.get_lmt_settings()
    list_missing_texts(ltm_settings[0], xlsxan_settings[0], ltm_settings[1], ltm_settings[2])
